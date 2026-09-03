"""Tests for the ptm_simple theme-first process (kept off the PTM pipeline)."""

from __future__ import annotations

import json
from datetime import date, timedelta

import pytest

from ptm_simple.radar import _member_snapshot, theme_radar


# ---------------------------------------------------------------- fixtures
@pytest.fixture
def watchlist_xlsx(tmp_path):
    """A miniature Master watchlist with the sheet's quirks reproduced:
    label in col J, a ticker sitting IN the label column, a section header,
    and a two-row theme."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Master watchlist"
    ws["C1"] = "AI HARDWARE"  # section header: ignored
    ws["H2"], ws["I2"], ws["J2"], ws["K2"] = "ARM", "AMD", "AI Semiconductor   chips", "NVDA"
    ws["G3"], ws["H3"], ws["I3"] = "QCOM", "INTC", "MU"  # continuation row
    ws["H4"], ws["I4"], ws["J4"] = "DLR", "IRM", "Data centre REIT"  # ticker in label col
    ws["K4"] = "EQIX"
    ws["H5"], ws["J5"] = "WDC", "Storgae Hardware/Memory"  # alias normalisation
    ws["I5"], ws["J5"] = "STX", "Storgae Hardware/Memory"
    path = tmp_path / "pack.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def members():
    """Two covered members: one aligned long, one divergent (short case)."""
    ref = date(2026, 9, 1)
    aligned = {
        "ticker": "AAA", "covered": True, "rev90": 4.0, "up30": 3, "down30": 0,
        "earnings_date": "2026-09-10", "days_to_print": 9,
        "revenue": 1e9, "net_income": 1e8, "cash": 5e8, "market_cap": 8e9,
    }
    divergent = {
        "ticker": "BBB", "covered": True, "rev90": -3.0, "up30": 0, "down30": 2,
        "earnings_date": "2026-09-12", "days_to_print": 11,
        "revenue": 5e8, "net_income": -2e7, "cash": -1e8, "market_cap": 4e8,
    }
    thin = {"ticker": "CCC", "covered": False, "rev90": None, "up30": None, "down30": None,
            "earnings_date": None, "days_to_print": None, "revenue": None, "net_income": None,
            "cash": None, "market_cap": None}
    return [aligned, divergent, thin], ref


def _row(members_list, **kw):
    base = {
        "theme": "Test theme", "thesis": "", "status": "ACTIVE", "lean": "long",
        "breadth": 0.5, "coverage": 0.67, "members_covered": 2, "members_total": 3,
        "prints_14d": [], "bellwether": "AAA", "divergent": ["BBB"], "members": members_list,
        "why_now": None,
    }
    base.update(kw)
    return base


# ---------------------------------------------------------------- thememap
def test_parse_watchlist_layout_and_aliases(watchlist_xlsx):
    from ptm_simple.thememap import parse_watchlist

    themes = {t["theme"]: sorted(t["members"]) for t in parse_watchlist(watchlist_xlsx, min_members=2)}
    # label alias applied; the ticker sitting in the label column is still a member
    assert themes["Data centre REIT"] == ["DLR", "EQIX", "IRM"]
    assert themes["Storage Hardware/Memory"] == ["STX", "WDC"]
    assert themes["AI Semiconductor chips"] == ["AMD", "ARM", "INTC", "MU", "NVDA", "QCOM"]
    # the section header never becomes a theme
    assert "AI HARDWARE" not in themes
    # continuation rows append to the current theme
    assert "MU" in themes["AI Semiconductor chips"]


def test_build_theme_map_reverse_index(watchlist_xlsx, tmp_path, monkeypatch):
    from ptm.config import set_roots
    from ptm_simple.thememap import build_theme_map

    set_roots(data=tmp_path, ideas=tmp_path / "ideas")
    out = build_theme_map(watchlist_xlsx)
    assert out["theme_count"] == 3
    assert set(out["ticker_themes"]["EQIX"]) == {"Data centre REIT"}


# ---------------------------------------------------------------- radar
def test_member_snapshot_reads_cache_and_age(tmp_path, monkeypatch):
    from ptm.config import set_roots

    set_roots(data=tmp_path, ideas=tmp_path / "ideas")
    fresh = tmp_path / "raw" / "expectations"
    fresh.mkdir(parents=True)
    payload = {
        "ticker": "AAA", "earnings_date": "2026-09-10",
        "revisions": {"available": True, "change_90d_pct": 4.0, "analysts_up_30d": 3, "analysts_down_30d": 0},
    }
    cache = fresh / "AAA.json"
    cache.write_text(json.dumps(payload))
    snap = _member_snapshot("AAA", {}, date(2026, 9, 1))
    assert snap["covered"] and snap["rev90"] == 4.0 and snap["days_to_print"] == 9

    stale = fresh / "BBB.json"
    stale.write_text(json.dumps(payload))
    import os
    from datetime import datetime

    old = datetime(2026, 6, 3).timestamp()
    os.utime(stale, (old, old))
    assert _member_snapshot("BBB", {}, date(2026, 9, 1))["covered"] is False


def _snap(ticker, rev90, dprint=None):
    return {"ticker": ticker, "covered": True, "rev90": rev90, "up30": 1, "down30": 0,
            "earnings_date": None, "days_to_print": dprint, "revenue": None,
            "net_income": None, "cash": None, "market_cap": None}


def test_theme_radar_status_thresholds(members):
    _, ref = members
    # one covered member, all up -> breadth +1, coverage 1.0
    row = theme_radar(_row([_snap("AAA", 4.0)]), {}, ref)
    assert row["status"] == "ACTIVE" and row["lean"] == "long" and row["breadth"] == 1.0
    # 2 up, 1 down of 5 covered -> breadth 0.2 -> WARM
    warm_members = [_snap("A", 4.0), _snap("B", 4.0), _snap("C", -4.0), _snap("D", 0.1), _snap("E", 0.2)]
    assert theme_radar(_row(warm_members), {}, ref)["status"] == "WARM"
    # one mildly up member -> breadth 0 -> COLD
    assert theme_radar(_row([_snap("A", 0.1)]), {}, ref)["status"] == "COLD"
    # divergent members are exactly those revising against the breadth sign
    row = theme_radar(_row([_snap("A", 4.0), _snap("B", -3.0)]), {}, ref)
    assert row["breadth"] == 0.0 and row["status"] == "COLD"  # netted flat
    # 3 up, 2 down of 5 -> breadth +0.2; the two downs diverge against the long lean
    row = theme_radar(_row([_snap("A", 4.0), _snap("B", 4.0), _snap("C", 4.0),
                            _snap("B2", -3.0), _snap("B3", -1.0)]), {}, ref)
    assert row["divergent"] == ["B2", "B3"]
    row = theme_radar(_row([_snap("A", -4.0), _snap("B", -4.0), _snap("S", 3.0)]), {}, ref)
    assert row["lean"] == "short" and row["divergent"] == ["S"]


# ---------------------------------------------------------------- select
# ---------------------------------------------------------------- gate
# ---------------------------------------------------------------- book
# ---------------------------------------------------------------- viewer glue
def _seed_simple_artifacts(isolate_roots):
    """Write a small set of simple-process artifacts under the isolated roots."""
    from ptm.config import data_dir, ideas_dir
    from ptm.io import write_json

    sdir = data_dir("simple")
    write_json(sdir / "theme_map.json", {"source": "xlsx", "theme_count": 3, "ticker_count": 5,
                                         "themes": [{"theme": "Data centre REIT", "members": ["DLR", "EQIX"]}]})
    write_json(sdir / "theme_map_wiki.json", {"source": "wikipedia-industry", "theme_count": 9,
                                              "ticker_count": 20, "wiki_fallbacks": 4, "themes": []})
    write_json(sdir / "radar_2026-09-01.json", {"as_of": "2026-09-01", "themes": []})
    write_json(sdir / "radar_2026-09-08.json", {"as_of": "2026-09-08", "themes": []})
    rdir = ideas_dir("simple", "Data centre REIT")
    rdir.mkdir(parents=True, exist_ok=True)
    (rdir / "DLR.md").write_text("# DLR long", encoding="utf-8")
    return sdir


def test_viewer_artifacts_inventory(isolate_roots):
    from ptm_simple import viewer

    _seed_simple_artifacts(isolate_roots)
    arts = viewer.artifacts()
    assert arts["maps"]["manual"]["exists"] and arts["maps"]["manual"]["themes"] == 3
    assert arts["maps"]["wiki"]["fallbacks"] == 4
    assert arts["radar_files"] == ["2026-09-01", "2026-09-08"] and arts["radar_date"] == "2026-09-08"
    assert "book" not in arts and "watchlist" not in arts
    assert arts["reports"] == 1


def test_viewer_theme_map_and_radar_readers(isolate_roots):
    from ptm_simple import viewer

    _seed_simple_artifacts(isolate_roots)
    got = viewer.get_theme_map("wiki")
    assert got["map"]["theme_count"] == 9 and got["source"] == "wiki"
    assert "error" in viewer.get_theme_map("bogus")
    radar = viewer.get_radar()  # latest wins
    assert radar["radar"]["as_of"] == "2026-09-08"
    dated = viewer.get_radar("2026-09-01")
    assert dated["radar"]["as_of"] == "2026-09-01"
    assert "error" in viewer.get_radar("2001-01-01")


def test_viewer_read_report_blocks_traversal(isolate_roots):
    from ptm_simple import viewer

    _seed_simple_artifacts(isolate_roots)
    ok = viewer.read_report("Data centre REIT/DLR.md")
    assert ok["markdown"].startswith("# DLR")
    assert viewer.read_report("../../curated/universe.csv") == {"error": "not found"}
    assert viewer.read_report("nope/missing.md") == {"error": "not found"}


def test_viewer_start_guards(isolate_roots):
    from ptm_simple import viewer

    ok, payload, code = viewer.start("nonsense", {})
    assert not ok and code == 400
    # analyze-all dives tickers, so it refuses to start on other LLM work
    ok, payload, code = viewer.start("analyze-all", {"theme": "Data centre REIT"}, other_work_running=True)
    assert not ok and code == 409 and "deep-dive batch" in payload["error"]
    assert not viewer.status()["running"]  # nothing was actually started


def test_viewer_theme_detail_unknown_theme(isolate_roots):
    from ptm_simple import viewer

    _seed_simple_artifacts(isolate_roots)
    assert "error" in viewer.get_theme_detail("Ghost theme", "manual")

# ---------------------------------------------------------------- wiki map
def test_wiki_map_keeps_sub_three_industries(isolate_roots, monkeypatch):
    """A Wikipedia industry with fewer than three names is still a theme: it is
    judged in isolation on absolute fundamentals, not dropped for thin company."""
    from ptm_simple import wiki_themes

    monkeypatch.setattr(wiki_themes, "_names",
                        lambda: {"AAA": "Alpha", "BBB": "Beta", "CCC": "Gamma"})
    monkeypatch.setattr(wiki_themes, "_yf_industries",
                        lambda: {"AAA": "", "BBB": "", "CCC": ""})
    monkeypatch.setattr(
        wiki_themes, "wiki_industries",
        lambda tickers, names, sleep_s=0.1: {
            "AAA": {"industries": ["Semiconductor Equipment"], "source": "wiki-infobox"},
            "BBB": {"industries": ["Semiconductor Equipment"], "source": "wiki-infobox"},
            "CCC": {"industries": ["Car Dealership"], "source": "wiki-infobox"},
        },
    )
    out = wiki_themes.build_theme_map_wiki()
    names = {t["theme"] for t in out["themes"]}
    assert "Car Dealership" in names and out["min_members"] == 1
    assert out["ticker_count"] == 3

    tight = wiki_themes.build_theme_map_wiki(min_members=3)
    assert "Car Dealership" not in {t["theme"] for t in tight["themes"]}
    assert tight["min_members"] == 3


def test_quant_flag_sole_member_gets_its_own_reason(isolate_roots):
    """A singleton theme has no peer median: the flag stays n/a but says why —
    judged on absolute multiples, not silently 'too few members'."""
    from ptm_simple.quant import _flag_rows

    row = {"theme": "car dealership", "pe1": 11.5, "peg1": 2.0}
    _flag_rows([row])
    assert row["flag"] == "n/a"
    assert "sole member" in row["flag_detail"]

    pair = [{"theme": "pair", "pe1": 10.0}, {"theme": "pair"}]
    _flag_rows(pair)
    assert pair[0]["flag"] == "n/a" and "too few members" in pair[0]["flag_detail"]


def test_simple_universe_falls_back_to_index_meta(isolate_roots):
    """A rebuilt fundamentals row has no sector/industry of its own (EDGAR does
    not carry them), so simple_universe fills the gap from the curated index
    table — otherwise the wiki map's ticker walk silently shrinks to the names
    an earlier cache still covered."""
    import pandas as pd
    from ptm.config import data_dir
    from ptm_simple.refresh import simple_universe

    fund_path = data_dir("curated", "yahoo_fundamentals.csv")
    fund_path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([{"ticker": "AAA", "name": "Alpha", "sector": "", "industry": ""}]).to_csv(
        fund_path, index=False)
    uni_path = data_dir("curated", "universe.csv")
    pd.DataFrame([{"ticker": "AAA", "name": "Alpha", "sector": "Industrials",
                   "industry": "Aerospace & Defense"}]).to_csv(uni_path, index=False)

    frame = simple_universe({"themes": [{"theme": "t", "members": ["AAA"]}]} )
    assert frame.iloc[0]["industry"] == "Aerospace & Defense"
    assert frame.iloc[0]["sector"] == "Industrials"


def test_wiki_industry_walk_fills_blank_industries(isolate_roots):
    """A fundamentals table rebuilt EDGAR-first carries no industry for its new
    rows; the map's ticker walk must still see those names, filling their
    classification from the curated index table instead of silently dropping
    them (that silent drop is how a 1295-ticker walk became 62)."""
    import pandas as pd
    from ptm.config import data_dir
    from ptm_simple import wiki_themes

    fund = data_dir("curated", "yahoo_fundamentals.csv")
    fund.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([
        {"ticker": "AAA", "name": "Alpha", "industry": "Aerospace & Defense"},
        {"ticker": "BBB", "name": "Beta", "industry": None},
    ]).to_csv(fund, index=False)
    uni = data_dir("curated", "universe.csv")
    pd.DataFrame([
        {"ticker": "AAA", "sector": "Industrials", "industry": "Aerospace & Defense"},
        {"ticker": "BBB", "sector": "Consumer Discretionary", "industry": "Car Dealership"},
    ]).to_csv(uni, index=False)

    out = wiki_themes._yf_industries()
    assert out["AAA"] == "Aerospace & Defense"
    assert out["BBB"] == "Car Dealership"


def test_wiki_map_dedupes_multi_label_tickers(isolate_roots, monkeypatch):
    """Wikidata labels a broad company with several industries (Salesforce
    carries nine); one membership per ticker keeps it only in its LARGEST
    theme, so a name is never ranked against the same fundamentals twice."""
    from ptm_simple import wiki_themes

    themes = [
        {"theme": "Software Industry", "members": ["CRM", "ORCL", "AAPL"]},
        {"theme": "Analytics", "members": ["CRM", "MSTR"]},
        {"theme": "Automation", "members": ["CRM"]},
    ]
    kept = {t["theme"]: t["members"] for t in wiki_themes.dedupe_memberships(themes)}
    assert kept["Software Industry"] == ["AAPL", "CRM", "ORCL"]  # CRM's largest label
    assert kept["Analytics"] == ["MSTR"]                          # CRM left, MSTR stays
    assert "Automation" not in kept                                # emptied and dropped

    # a map built through the same path carries no duplicate membership
    monkeypatch.setattr(wiki_themes, "_names", lambda: {"AAA": "A", "BBB": "B", "CCC": "C"})
    monkeypatch.setattr(wiki_themes, "_yf_industries", lambda: {"AAA": "", "BBB": "", "CCC": ""})
    monkeypatch.setattr(
        wiki_themes, "wiki_industries",
        lambda tickers, names, sleep_s=0.1: {
            "AAA": {"industries": ["Big Industry"], "source": "wiki-infobox"},
            "BBB": {"industries": ["Big Industry", "Small Niche"], "source": "wiki-infobox"},
            "CCC": {"industries": ["Small Niche"], "source": "wiki-infobox"},
        },
    )
    out = wiki_themes.build_theme_map_wiki()
    assert out["ticker_themes"]["BBB"] == ["Big Industry"]
    assert out["ticker_themes"]["CCC"] == ["Small Niche"]
